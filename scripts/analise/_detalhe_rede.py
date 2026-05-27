"""Detalhamento de uma rede: lista completa de lojas presentes/ausentes."""
import sys, io, os, re, unicodedata
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

REDE = sys.argv[1] if len(sys.argv) > 1 else 'PREZUNIC'

def norm_placa(p):
    if not isinstance(p, str): return ''
    return p.upper().replace('-', '').replace(' ', '')

def fmt(v):
    if v is None: return ''
    if hasattr(v, 'strftime'): return v.strftime('%H:%M')
    return str(v)

def loja_key(s):
    if not isinstance(s, str): return None
    m = re.search(r'(?:Loja|Filial)\s*(\d{1,3})', s, re.IGNORECASE)
    if m: return int(m.group(1))
    return None

def le(path, prefer):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = None
    for nome in [prefer, '25', '25.05']:
        if nome in wb.sheetnames:
            sheet = wb[nome]; break
    if sheet is None: sheet = wb.active
    rows = []
    for r in range(5, sheet.max_row + 1):
        loja = sheet.cell(r, 1).value
        if not loja: continue
        rows.append({
            'loja_raw': str(loja).strip(),
            'loja_key': loja_key(loja),
            'mot': str(sheet.cell(r, 2).value or '').strip()[:18],
            'cod': str(sheet.cell(r, 3).value or '').strip(),
            'placa': norm_placa(sheet.cell(r, 4).value),
        })
    return rows

manual = le(f'docs/conversas-tia-erica/dia-25/kpis-manuais/KPI-{REDE}-MANUAL.xlsx', '25')
sistema = le(f'docs/conversas-tia-erica/dia-25/kpis-gerados/KPI-{REDE}-2026-05-25.xlsx', '25.05')

md = {r['loja_key']: r for r in manual if r['loja_key'] is not None}
sd = {r['loja_key']: r for r in sistema if r['loja_key'] is not None}

print(f"=== {REDE} ===  Manual={len(manual)} lojas | Sistema={len(sistema)} lojas\n")
print(f"{'Key':<5} {'Loja Manual':<42} | {'M mot/cod/placa':<32} | {'S mot/cod/placa':<32}")
print('-'*120)

chaves = sorted(set(list(md.keys()) + list(sd.keys())))
for k in chaves:
    m = md.get(k); s = sd.get(k)
    loja = (m or s)['loja_raw'][:40]
    m_str = f"{m['mot']}/{m['cod']}/{m['placa']}" if m else 'AUSENTE NO MANUAL'
    s_str = f"{s['mot']}/{s['cod']}/{s['placa']}" if s else 'AUSENTE NO SISTEMA'
    marker = ''
    if not s: marker = ' << FALTA NO SISTEMA'
    elif not m: marker = ' << SOBROU NO SISTEMA'
    print(f"{k:<5} {loja:<42} | {m_str:<32} | {s_str:<32}{marker}")
