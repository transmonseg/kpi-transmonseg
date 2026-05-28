"""Taxa de acerto ZS dia 19 com tolerancia de 5min."""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

with open('docs/auditoria/dump-dia-19/06-unitrac.md', encoding='utf-8') as f:
    md = f.read()

blocos = re.split(r'\n## ([A-Z0-9-]+)\s*\(', md)
paradas_por_placa = {}
for i in range(1, len(blocos), 2):
    placa = blocos[i].replace('-', '').upper()
    conteudo = blocos[i+1]
    lojas = []
    for m in re.finditer(r'\|\s*(LOJA|FORA_BASE)\s*\|\s*(\d{2}:\d{2})-(\d{2}:\d{2})\s*\|\s*(\d+)min\s*\|\s*(\S+)\s*\|\s*([^|]+)\|', conteudo):
        lojas.append({
            'tipo': m.group(1), 'chd': m.group(2), 'sai': m.group(3),
            'dur': int(m.group(4)), 'cod_loja': m.group(5).strip(),
            'nome': m.group(6).strip(),
        })
    paradas_por_placa[placa] = lojas

def hhmm_to_min(s):
    h, m = s.split(':')
    return int(h) * 60 + int(m)

wb = openpyxl.load_workbook('docs/conversas-tia-erica/dia-19/kpis-sistema/KPI-ZONA_SUL-2026-05-19 (9).xlsx', data_only=True)
ws = wb['19.05']

def fmt(v):
    if v is None or v == '': return None
    if hasattr(v, 'strftime'): return v.strftime('%H:%M')
    return None

def num_loja(nome):
    if not nome: return None
    m = re.search(r'Loja\s+(\d+)', str(nome), re.IGNORECASE)
    if m: return m.group(1)
    return None

TOL = 5
total = 0
acerto_completo = 0
acerto_chd_so = 0
sem_rastre_ok = 0
sem_rastre_errado = 0
inventado = 0
loja_errada = 0
sem_ts = 0
problemas = []

for r in range(5, ws.max_row + 1):
    loja_raw = ws.cell(r, 1).value
    if not loja_raw: continue
    total += 1
    placa = (ws.cell(r, 4).value or '').replace('-', '').upper()
    chd_v = ws.cell(r, 6).value
    sai_lj_v = ws.cell(r, 7).value
    chd = fmt(chd_v)
    sai = fmt(sai_lj_v)

    chd_str = str(chd_v or '')
    is_sem_rast = 'SEM' in chd_str.upper() or 'RASTRE' in chd_str.upper()
    is_sem_ts = (not chd or not sai) and not is_sem_rast

    lojas = paradas_por_placa.get(placa, [])
    num = num_loja(loja_raw)

    if is_sem_rast:
        if not [p for p in lojas if p['tipo'] == 'LOJA']:
            sem_rastre_ok += 1
        else:
            sem_rastre_errado += 1
            problemas.append((loja_raw, placa, 'SEM-RASTRE mas Unitrac tem LOJA'))
        continue

    if is_sem_ts:
        sem_ts += 1
        problemas.append((loja_raw, placa, 'sistema sem timestamp'))
        continue

    chd_m = hhmm_to_min(chd)
    achou_loja_correta = None
    achou_qualquer_match = None
    for p in lojas:
        if p['tipo'] != 'LOJA': continue
        p_chd = hhmm_to_min(p['chd'])
        if abs(p_chd - chd_m) > TOL: continue
        achou_qualquer_match = p
        if num and num in p['nome']:
            achou_loja_correta = p
            break

    if achou_loja_correta:
        if abs(hhmm_to_min(achou_loja_correta['sai']) - hhmm_to_min(sai)) <= TOL:
            acerto_completo += 1
        else:
            acerto_chd_so += 1
            problemas.append((loja_raw, placa, f'saida sistema={sai} Unitrac={achou_loja_correta["sai"]}'))
    elif achou_qualquer_match:
        loja_errada += 1
        problemas.append((loja_raw, placa, f'GPS em "{achou_qualquer_match["nome"][:25]}" {achou_qualquer_match["chd"]}-{achou_qualquer_match["sai"]}'))
    else:
        inventado += 1
        problemas.append((loja_raw, placa, f'sem GPS perto de {chd}'))

print(f"Total ZS dia 19: {total}")
print(f"  Acerto completo (chd+sai ±5min):  {acerto_completo}")
print(f"  Acerto so chegada (saida +5min):  {acerto_chd_so}")
print(f"  SEM-RASTRE correto:               {sem_rastre_ok}")
print(f"  SEM-RASTRE errado:                {sem_rastre_errado}")
print(f"  Sem timestamp gerado:             {sem_ts}")
print(f"  Loja errada (match em outra):     {loja_errada}")
print(f"  Inventado (sem GPS perto):        {inventado}")

acertos_estrito = acerto_completo + sem_rastre_ok
acertos_tolerante = acerto_completo + acerto_chd_so + sem_rastre_ok
print(f"\n  Acerto estrito: {acertos_estrito}/{total} = {acertos_estrito*100//total}%")
print(f"  Acerto tolerante (saida ±5min):  {acertos_tolerante}/{total} = {acertos_tolerante*100//total}%")

print(f"\n  PROBLEMAS:")
for loja, placa, motivo in problemas:
    print(f"    {str(loja)[:32]:<32} {placa:<10} {motivo}")
