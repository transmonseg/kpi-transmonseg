"""Compara KPI sistema vs Unitrac PDF (parser do projeto) por rede dia 19.
Tolerancia 10min. SEM-RASTRE valido = OK. INVENTADO + LOJA_ERRADA sao bugs.
Gera docs/conversas-tia-erica/dia-19/analise-{REDE}.md
"""
import sys, io, re, json, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

with open('docs/conversas-tia-erica/dia-19/unitrac-pdf.json', encoding='utf-8') as f:
    UNITRAC = json.load(f)

TOL = 10

def hhmm_to_min(s):
    if not s or ':' not in str(s): return None
    h, m = s.split(':')
    return int(h)*60 + int(m)

def fmt(v):
    if v is None or v == '': return None
    if hasattr(v, 'strftime'): return v.strftime('%H:%M')
    return None

def num_loja(nome):
    if not nome: return None
    # Princesa, Carrefour: pode usar nome literal
    m = re.search(r'\bLoja\s+(\d+)', str(nome), re.IGNORECASE)
    if m: return m.group(1).lstrip('0')
    m = re.search(r'\bFilial\s+(\d+)', str(nome), re.IGNORECASE)
    if m: return m.group(1).lstrip('0')
    return None

def cod_in_paradas(num, paradas):
    """Verifica se algum cod_loja de paradas LOJA contem o numero da loja."""
    if not num: return False
    for p in paradas:
        if p['tipo'] != 'LOJA': continue
        cod = str(p.get('cod_loja') or '')
        nome = str(p.get('nome') or '').upper()
        # match por cod_loja (varias formas)
        if num == cod: return True
        if cod.endswith(num): return True
        if num in cod: return True
        # match por nome (ex "43 - ZONA SUL - BARRA")
        if re.search(rf'\b{num}\b', nome): return True
    return False

def analisar_rede(rede_nome, kpi_path):
    if not os.path.exists(kpi_path):
        return None
    wb = openpyxl.load_workbook(kpi_path, data_only=True)
    # Procura aba 19.05
    ws = None
    for nome in ['19.05', '19']:
        if nome in wb.sheetnames:
            ws = wb[nome]; break
    if ws is None:
        ws = wb.active

    total = 0
    acerto_completo = 0
    acerto_chd_so = 0
    saida_off = []  # lista de (loja, placa, sis_sai, real_sai)
    sem_rastre_ok = 0
    sem_rastre_errado = []
    inventado = []
    loja_errada = []  # (loja, placa, real_nome, real_chd-real_sai)
    sem_ts = 0
    todas_linhas = []

    for r in range(5, ws.max_row + 1):
        loja_raw = ws.cell(r, 1).value
        if not loja_raw: continue
        total += 1
        placa_raw = ws.cell(r, 4).value
        placa = (str(placa_raw) if placa_raw else '').replace('-', '').upper().strip()
        chd_v = ws.cell(r, 6).value
        sai_v = ws.cell(r, 7).value
        chd = fmt(chd_v)
        sai = fmt(sai_v)
        chd_str = str(chd_v or '')
        is_sem_rast = 'SEM' in chd_str.upper() and ('RAST' in chd_str.upper() or 'RAS' in chd_str.upper())
        is_sem_ts = (not chd or not sai) and not is_sem_rast

        paradas = UNITRAC.get(placa, [])
        lojas_un_raw = [p for p in paradas if p['tipo'] == 'LOJA']
        # Consolida paradas LOJA com mesmo cod_loja e gap <=15min: 05:05-06:05 + 06:06-12:31 → 05:05-12:31
        lojas_un = []
        for p in lojas_un_raw:
            if lojas_un and lojas_un[-1].get('cod_loja') == p.get('cod_loja') and lojas_un[-1].get('cod_loja') not in (None, '', '-'):
                gap = hhmm_to_min(p['chd']) - hhmm_to_min(lojas_un[-1]['sai'])
                if gap is not None and gap <= 15:
                    lojas_un[-1] = {**lojas_un[-1], 'sai': p['sai']}
                    continue
            lojas_un.append(p)
        num = num_loja(str(loja_raw))

        if is_sem_rast:
            if not lojas_un:
                sem_rastre_ok += 1
            else:
                sem_rastre_errado.append((str(loja_raw)[:35], placa, len(lojas_un)))
            continue

        if is_sem_ts:
            sem_ts += 1
            continue

        chd_m = hhmm_to_min(chd)
        achou_correto = None
        achou_outro = None
        for p in lojas_un:
            p_chd = hhmm_to_min(p['chd'])
            if p_chd is None or abs(p_chd - chd_m) > TOL: continue
            achou_outro = p
            cod_p = str(p.get('cod_loja') or '')
            nome_p = str(p.get('nome') or '').upper()
            if num and (num == cod_p or cod_p.endswith(num) or num in cod_p or re.search(rf'\b{num}\b', nome_p)):
                achou_correto = p
                break

        if achou_correto:
            sai_real = hhmm_to_min(achou_correto['sai'])
            sai_sist = hhmm_to_min(sai)
            if sai_real is not None and sai_sist is not None and abs(sai_real - sai_sist) <= TOL:
                acerto_completo += 1
            else:
                acerto_chd_so += 1
                saida_off.append((str(loja_raw)[:35], placa, sai, achou_correto['sai']))
        elif achou_outro:
            # Quando num é None (rede sem código) ou placa só serve 1 loja: provavel match
            # Verificar por tokens fortes do nome
            STOP = {'LOJA', 'FILIAL', 'PREZUNIC', 'ASSAI', 'CARREFOUR', 'PRINCESA',
                    'ATACADAO', 'SUPERPRIX', 'GUANABARA', 'ZONA', 'SUL', 'ARMAZEM',
                    'PAX', 'FEIRA', 'NOVA', 'SAMS', 'CLUB', 'CAB', 'PETROPOLIS',
                    'SUPER', 'EMANUEL', 'VIANENSE', 'MUNDIAL', 'SENDAS'}
            def toks(s):
                import unicodedata
                s = unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode('ascii')
                return {t for t in re.findall(r'[A-Z]{4,}', s.upper()) if t not in STOP}
            t_sis = toks(str(loja_raw))
            t_real = toks(str(achou_outro.get('nome') or ''))
            if t_sis & t_real:
                # token forte em comum → mesma loja
                sai_real = hhmm_to_min(achou_outro['sai'])
                sai_sist = hhmm_to_min(sai)
                if sai_real is not None and sai_sist is not None and abs(sai_real - sai_sist) <= TOL:
                    acerto_completo += 1
                else:
                    acerto_chd_so += 1
                    saida_off.append((str(loja_raw)[:35], placa, sai, achou_outro['sai']))
            elif num is None and len(lojas_un) == 1:
                # Placa fez UMA loja só, sistema atribuiu sem token compartilhado.
                # Considera match (lacuna no nome).
                sai_real = hhmm_to_min(achou_outro['sai'])
                sai_sist = hhmm_to_min(sai)
                if sai_real is not None and sai_sist is not None and abs(sai_real - sai_sist) <= TOL:
                    acerto_completo += 1
                else:
                    acerto_chd_so += 1
                    saida_off.append((str(loja_raw)[:35], placa, sai, achou_outro['sai']))
            else:
                loja_errada.append((str(loja_raw)[:35], placa,
                                   f"{achou_outro['chd']}-{achou_outro['sai']}",
                                   (achou_outro.get('nome') or '')[:35]))
        else:
            # placa não tem parada perto, mas pode estar em FORA_BASE
            fora_match = next((p for p in paradas if p['tipo'] == 'FORA_BASE'
                               and hhmm_to_min(p['chd']) is not None
                               and abs(hhmm_to_min(p['chd']) - chd_m) <= TOL), None)
            if fora_match:
                # Sistema atribuiu loja em parada FORA_BASE — pode ser que sistema acertou (loja sem geofence)
                # Mas se nao tem LOJA classificada, consideramos "inventado parcial"
                inventado.append((str(loja_raw)[:35], placa, f"FORA_BASE perto {fora_match['chd']}"))
            else:
                inventado.append((str(loja_raw)[:35], placa, f"sem parada perto de {chd}"))

    return {
        'total': total,
        'acerto_completo': acerto_completo,
        'acerto_chd_so': acerto_chd_so,
        'saida_off': saida_off,
        'sem_rastre_ok': sem_rastre_ok,
        'sem_rastre_errado': sem_rastre_errado,
        'inventado': inventado,
        'loja_errada': loja_errada,
        'sem_ts': sem_ts,
    }

REDES = [
    ('ZONA_SUL', 'KPI-ZONA_SUL-2026-05-19 (9).xlsx'),
    ('ASSAI', 'KPI-ASSAI-2026-05-19 (4).xlsx'),
    ('ATACADAO', 'KPI-ATACADAO-2026-05-19 (5).xlsx'),
    ('CARREFOUR', 'KPI-CARREFOUR-2026-05-19 (4).xlsx'),
    ('PREZUNIC', 'KPI-PREZUNIC-2026-05-19 (4).xlsx'),
    ('PRINCESA', 'KPI-PRINCESA-2026-05-19 (11).xlsx'),
    ('SUPERPRIX', 'KPI-SUPERPRIX-2026-05-19 (4).xlsx'),
    ('SUPERCOMPRAS', 'KPI-SUPERCOMPRAS-2026-05-19 (2).xlsx'),
    ('ARMAZEM_GRAO', 'KPI-ARMAZEM_GRAO-2026-05-19 (9).xlsx'),
    ('GUANABARA', 'KPI-GUANABARA-2026-05-19 (8).xlsx'),
    ('MUNDIAL', 'KPI-MUNDIAL-2026-05-19 (4).xlsx'),
    ('SUPER_PAX', 'KPI-SUPER_PAX-2026-05-19 (3).xlsx'),
    ('FEIRA_NOVA', 'KPI-FEIRA_NOVA-2026-05-19 (4).xlsx'),
    ('CAB_PETROPOLIS', 'KPI-CAB_PETROPOLIS-2026-05-19 (3).xlsx'),
    ('SAMS_CLUB', 'KPI-SAMS_CLUB-2026-05-19 (3).xlsx'),
    ('SENDAS', 'KPI-SENDAS-2026-05-19 (5).xlsx'),
    ('VIANENSE', 'KPI-VIANENSE-2026-05-19 (3).xlsx'),
]

resultados = {}
print(f"{'Rede':<18} | {'Total':<5} | {'OK':<4} | {'Sai off':<7} | {'SR':<4} | {'Vazio':<5} | {'Loja err':<8} | {'Inventado':<10} | {'Taxa'}")
print('-' * 120)
for rede, arq in REDES:
    path = f'docs/conversas-tia-erica/dia-19/kpis-sistema/{arq}'
    res = analisar_rede(rede, path)
    if res is None:
        print(f"{rede:<18} | (sem KPI)")
        continue
    resultados[rede] = res
    acerto = res['acerto_completo'] + res['acerto_chd_so'] + res['sem_rastre_ok']
    taxa = acerto * 100 // res['total'] if res['total'] else 0
    print(f"{rede:<18} | {res['total']:<5} | {res['acerto_completo']:<4} | {res['acerto_chd_so']:<7} | {res['sem_rastre_ok']:<4} | {res['sem_ts']:<5} | {len(res['loja_errada']):<8} | {len(res['inventado']):<10} | {taxa}%")

    # Salva MD por rede
    md = [f'# Analise {rede} — dia 19/05/2026 (tol 10min)', '']
    md.append(f"**Total lojas:** {res['total']}")
    md.append(f"- Acerto completo: {res['acerto_completo']}")
    md.append(f"- Chegada certa, saida >10min: {res['acerto_chd_so']}")
    md.append(f"- SEM-RASTRE valido: {res['sem_rastre_ok']}")
    md.append(f"- Sem timestamp: {res['sem_ts']}")
    md.append(f"- **LOJA ERRADA (bug):** {len(res['loja_errada'])}")
    md.append(f"- **INVENTADO (bug):** {len(res['inventado'])}")
    md.append(f"- SEM-RASTRE errado: {len(res['sem_rastre_errado'])}")
    md.append('')
    if res['saida_off']:
        md.append('## Saidas erradas (chd OK, saida diverge >10min)')
        md.append('| Loja | Placa | Sai sistema | Sai Unitrac |')
        md.append('|---|---|---|---|')
        for loja, pl, ss, sr in res['saida_off']:
            md.append(f"| {loja} | {pl} | {ss} | {sr} |")
        md.append('')
    if res['loja_errada']:
        md.append('## Loja ERRADA (timestamp bate mas loja outra)')
        md.append('| Loja sistema | Placa | Horario | Loja real (Unitrac) |')
        md.append('|---|---|---|---|')
        for loja, pl, h, nome in res['loja_errada']:
            md.append(f"| {loja} | {pl} | {h} | {nome} |")
        md.append('')
    if res['inventado']:
        md.append('## INVENTADO (sistema deu timestamp sem GPS perto)')
        md.append('| Loja | Placa | Motivo |')
        md.append('|---|---|---|')
        for loja, pl, motivo in res['inventado']:
            md.append(f"| {loja} | {pl} | {motivo} |")
        md.append('')
    if res['sem_rastre_errado']:
        md.append('## SEM-RASTRE errado (Unitrac tinha GPS LOJA)')
        for loja, pl, n in res['sem_rastre_errado']:
            md.append(f"- {loja} {pl}: Unitrac tem {n} paradas LOJA")
        md.append('')

    with open(f'docs/conversas-tia-erica/dia-19/analise-{rede}.md', 'w', encoding='utf-8') as f:
        f.write('\n'.join(md))

# Resumo
resumo = ['# Resumo geral dia 19/05/2026', '']
resumo.append(f"| Rede | Total | OK | Sai off | SR | Vazio | Loja err | Inventado | Taxa |")
resumo.append(f"|---|---|---|---|---|---|---|---|---|")
total_geral = 0
acerto_geral = 0
bugs_geral = 0
for rede, res in resultados.items():
    acerto = res['acerto_completo'] + res['acerto_chd_so'] + res['sem_rastre_ok']
    taxa = acerto * 100 // res['total'] if res['total'] else 0
    bugs = len(res['loja_errada']) + len(res['inventado']) + len(res['sem_rastre_errado']) + len(res['saida_off'])
    resumo.append(f"| {rede} | {res['total']} | {res['acerto_completo']} | {res['acerto_chd_so']} | {res['sem_rastre_ok']} | {res['sem_ts']} | {len(res['loja_errada'])} | {len(res['inventado'])} | {taxa}% |")
    total_geral += res['total']
    acerto_geral += acerto
    bugs_geral += bugs

taxa_geral = acerto_geral * 100 // total_geral if total_geral else 0
resumo.append('')
resumo.append(f"**TOTAL:** {acerto_geral}/{total_geral} ({taxa_geral}%) | {bugs_geral} bugs (loja errada + inventado + saida off + sem-rastre errado)")

with open('docs/conversas-tia-erica/dia-19/resumo-geral.md', 'w', encoding='utf-8') as f:
    f.write('\n'.join(resumo))

print(f"\n>> docs/conversas-tia-erica/dia-19/resumo-geral.md")
print(f">> TOTAL: {acerto_geral}/{total_geral} = {taxa_geral}% | {bugs_geral} bugs")
