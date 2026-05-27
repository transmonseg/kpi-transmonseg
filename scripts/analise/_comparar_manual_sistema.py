"""Comparar KPI manual (gabarito tia Erica) vs KPI sistema (Vercel) dia 25/05."""
import sys, io, os, re, unicodedata
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

REDES = ['ASSAI', 'ATACADAO', 'CARREFOUR', 'PREZUNIC', 'PRINCESA', 'SUPERPRIX']

def norm_placa(p):
    if not isinstance(p, str): return ''
    return p.upper().replace('-', '').replace(' ', '')

def norm_nome(n):
    if not isinstance(n, str): return ''
    n = unicodedata.normalize('NFD', n).encode('ascii', 'ignore').decode('ascii')
    return n.upper().strip()

def fmt_hora(v):
    if v is None or v == '': return ''
    if hasattr(v, 'strftime'): return v.strftime('%H:%M')
    return str(v)

STOP = set(['LOJA', 'FILIAL', 'PREZUNIC', 'ASSAI', 'ATACADAO', 'CARREFOUR', 'PRINCESA',
            'SUPERPRIX', 'GUANABARA', 'SENDAS', 'VIANENSE', 'ZONA', 'SUL', 'ARMAZEM', 'MUNDIAL',
            'EMANUEL', 'PAX', 'FEIRA', 'NOVA', 'SAMS', 'CLUB', 'CAB', 'PETROPOLIS', 'SUPER',
            'ASSAÍ', 'I', 'II', 'III', 'IV', 'BARRA'])

def loja_key(s):
    """Tenta extrair número da loja. Se não houver, retorna nome tokenizado."""
    if not isinstance(s, str): return None
    m = re.search(r'(?:Loja|Filial)\s*(\d{1,3})', s, re.IGNORECASE)
    if m: return f"#{int(m.group(1))}"
    # Senão: tokens fortes do nome
    n = norm_nome(s).replace('-', ' ').replace('/', ' ')
    n = re.sub(r'[^A-Z0-9 ]', ' ', n)
    toks = [t for t in n.split() if len(t) >= 3 and t not in STOP]
    if not toks: return None
    return 'N:' + '-'.join(toks[:3])

def le(path, prefer):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = None
    for nome in [prefer, '25', '25.05']:
        if nome in wb.sheetnames:
            sheet = wb[nome]; break
    if sheet is None: sheet = wb.active
    rows = []
    for r in range(5, sheet.max_row + 1):
        loja = sheet.cell(r, 1).value
        if not loja: continue
        rows.append({
            'loja_raw': str(loja).strip(),
            'key': loja_key(loja),
            'mot': str(sheet.cell(r, 2).value or '').strip(),
            'cod': str(sheet.cell(r, 3).value or '').strip(),
            'placa': norm_placa(sheet.cell(r, 4).value),
            'saida_cd': fmt_hora(sheet.cell(r, 5).value),
            'chd': fmt_hora(sheet.cell(r, 6).value),
            'sl': fmt_hora(sheet.cell(r, 7).value),
        })
    return sheet.title, rows

def comparar(rede):
    mp = f'docs/conversas-tia-erica/dia-25/kpis-manuais/KPI-{rede}-MANUAL.xlsx'
    sp = f'docs/conversas-tia-erica/dia-25/kpis-gerados/KPI-{rede}-2026-05-25.xlsx'
    if not os.path.exists(mp) or not os.path.exists(sp): return None
    sm, manual = le(mp, '25')
    ss, sistema = le(sp, '25.05')

    # indexar
    md = {r['key']: r for r in manual if r['key'] is not None}
    sd = {r['key']: r for r in sistema if r['key'] is not None}
    chaves = sorted(set(list(md.keys()) + list(sd.keys())))

    divs = []
    for k in chaves:
        m = md.get(k); s = sd.get(k)
        if not m and s:
            divs.append({'key': k, 'tipo': 'EXTRA_SIS', 'manual': None, 'sistema': s})
            continue
        if m and not s:
            divs.append({'key': k, 'tipo': 'FALTA_SIS', 'manual': m, 'sistema': None})
            continue
        nome_m = norm_nome(m['mot']).split()[0] if m['mot'] else ''
        nome_s = norm_nome(s['mot']).split()[0] if s['mot'] else ''
        nome_diff = nome_m != nome_s and bool(nome_m) and bool(nome_s)
        placa_diff = m['placa'] != s['placa'] and bool(m['placa']) and bool(s['placa'])
        sem_mot = not s['mot'] and m['mot']
        if nome_diff or placa_diff or sem_mot:
            flags = []
            if nome_diff: flags.append('NOME')
            if placa_diff: flags.append('PLACA')
            if sem_mot: flags.append('VAZIO')
            divs.append({
                'key': k, 'tipo': '+'.join(flags), 'manual': m, 'sistema': s,
            })
    return {
        'manual_aba': sm, 'sistema_aba': ss,
        'manual_total': len(manual), 'sistema_total': len(sistema),
        'divs': divs,
    }

print(f"{'Rede':<12} | {'M aba':<8} | {'S aba':<8} | {'M':<5} | {'S':<5} | {'Diff':<5}")
print('-' * 70)
all_results = {}
for rede in REDES:
    r = comparar(rede)
    all_results[rede] = r
    if not r:
        print(f"{rede:<12} | (sem arquivo)")
        continue
    print(f"{rede:<12} | {r['manual_aba']:<8} | {r['sistema_aba']:<8} | {r['manual_total']:<5} | {r['sistema_total']:<5} | {len(r['divs']):<5}")

for rede in REDES:
    r = all_results.get(rede)
    if not r or not r['divs']: continue
    print(f"\n\n{'='*120}")
    print(f"  {rede}  ({len(r['divs'])} divergencias)")
    print('='*120)
    print(f"{'Key':<22} {'Tipo':<14} {'Manual mot/cod/placa':<40} {'Sistema mot/cod/placa':<40}")
    print('-' * 120)
    for d in r['divs']:
        m = d['manual']; s = d['sistema']
        if d['tipo'] == 'FALTA_SIS':
            mm = f"{m['mot'][:18]}/{m['cod']}/{m['placa']}"
            print(f"{d['key']:<22} {d['tipo']:<14} {mm:<40} (loja '{m['loja_raw'][:40]}' FALTA NO SISTEMA)")
        elif d['tipo'] == 'EXTRA_SIS':
            ss = f"{s['mot'][:18]}/{s['cod']}/{s['placa']}"
            print(f"{d['key']:<22} {d['tipo']:<14} (manual nao tem)                         {ss}")
        else:
            mm = f"{m['mot'][:18]}/{m['cod']}/{m['placa']}"
            ss = f"{s['mot'][:18]}/{s['cod']}/{s['placa']}"
            print(f"{d['key']:<22} {d['tipo']:<14} {mm:<40} {ss}")
