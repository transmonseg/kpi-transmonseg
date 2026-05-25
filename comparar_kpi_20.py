import openpyxl, datetime

def fmt(t):
    if t is None or t == '': return '-'
    if isinstance(t, datetime.time): return t.strftime('%H:%M')
    return str(t)

def to_min(t_str):
    s = str(t_str or '').strip()
    blocklist = ['-','SEM RASTREADOR','SEM','NAO FOI AO CLIENTE','SEM DADOS','#VALUE!','RASTREADOR','FOI  AO','CLIENTE']
    for x in blocklist:
        if x in s.upper(): return None
    if not s or s == 'None': return None
    try:
        p = s.split(':')
        return int(p[0])*60 + int(p[1])
    except: return None

# ===================== SISTEMA =====================
wb_s = openpyxl.load_workbook(r'C:\Users\media\Downloads\KPI-ZONA_SUL-2026-05-20 (1).xlsx', data_only=True)
ws_s = wb_s['20.05']
sis = {}
for i, row in enumerate(ws_s.iter_rows(values_only=True)):
    if i < 3: continue
    f = (row[0] or '').strip()
    if not f: continue
    saida_cd=row[4]; chd=row[5]; saida_loja=row[6]; tempo=row[13]
    mot = (row[1] or '').strip(); placa = (row[3] or '').strip()
    if saida_cd == 'SEM RASTREADOR': st = 'SEM RASTREADOR'
    elif not saida_cd: st = 'SEM DADOS'
    else: st = 'COM GPS'
    sis[f] = {
        'mot': mot, 'placa': placa,
        'saida_cd': fmt(saida_cd), 'chd': fmt(chd),
        'saida_loja': fmt(saida_loja), 'tempo': fmt(tempo),
        'status': st
    }

# ===================== MANUAL =====================
wb_m = openpyxl.load_workbook(r'C:\Users\media\Downloads\KPI_ZONA_SUL_manual_copy.xlsx', data_only=True)
ws_m = wb_m['20']
man = {}
for i, row in enumerate(ws_m.iter_rows(values_only=True)):
    if i < 2: continue
    f = (row[0] or '').strip()
    if not f: continue
    mot=(row[1] or '').strip(); placa=(row[2] or '').strip()
    s4=row[3]; s5=row[4]; s6=row[5]; s8=row[7]
    raw4 = str(s4 or '').upper().strip()
    if 'SEM' in raw4: st='SEM RASTREADOR'
    elif raw4.startswith('N') and len(raw4) < 5: st='NAO FOI AO CLIENTE'
    elif not s4: st='SEM DADOS'
    else: st='COM DADOS'
    e = {
        'mot': mot, 'placa': placa,
        'saida_cd': fmt(s4) if isinstance(s4, datetime.time) else str(s4 or '-'),
        'chd': fmt(s5) if isinstance(s5, datetime.time) else str(s5 or '-'),
        'saida_loja': fmt(s6) if isinstance(s6, datetime.time) else str(s6 or '-'),
        'tempo': fmt(s8) if isinstance(s8, datetime.time) else str(s8 or '-'),
        'status': st, 'row': i+1
    }
    if f not in man: man[f] = []
    man[f].append(e)

# ===================== PARES DE COMPARACAO =====================
pairs = [
    ('MEGA BOX 01 - Olaria',              'Zona Sul Mega Box 1 - Olaria'),
    ('MEGA BOX 02 - Olaria',              'Zona Sul Mega Box 2 - Recreio'),
    ('Zona Sul Loja 01 - Ipanema',        'Zona Sul Loja 01 - Ipanema'),
    ('Zona Sul Loja 03 - Copacabana I',   'Zona Sul Loja 03 - Copacabana I'),
    ('Zona Sul Loja 04 - Copacabana II',  'Zona Sul Loja 04 - Copacabana II'),
    ('Zona Sul Loja 05 - Copacabana III', 'Zona Sul Loja 05 - Copacabana III'),
    ('Zona Sul Loja 07 - Leblon',         'Zona Sul Loja 07 - Leblon'),
    ('Zona Sul Loja 08 - Ipanema',        'Zona Sul Loja 08 - Ipanema'),
    ('Zona Sul Loja 09 - Ipanema',        'Zona Sul Loja 09 - Ipanema'),
    ('Zona Sul Loja 10 - Recreio',        'Zona Sul Loja 10 - Recreio'),
    ('Zona Sul Loja 11 - Leblon',         'Zona Sul Loja 11 - Leblon'),
    ('Zona Sul Loja 12 - Leme',           'Zona Sul Loja 12 - Leme'),
    ('Zona Sul Loja 13 - Angra',          'Zona Sul Loja 13 - Angra'),
    ('Zona Sul Loja 14 - Leblon',         'Zona Sul Loja 14 - Leblon'),
    ('Zona Sul Loja 15 - Leblon',         'Zona Sul Loja 15 - Leblon'),
    ('Zona Sul Loja 17 - Barra',          'Zona Sul Loja 17 - Barra'),
    ('Zona Sul Loja 18 - Copacabana',     'Zona Sul Loja 18 - Copacabana'),
    ('Zona Sul Loja 19 - Copacabana',     'Zona Sul Loja 19 - Copacabana'),
    ('Zona Sul Loja 20 - Botafogo',       'Zona Sul Loja 20 - Botafogo'),
    ('Zona Sul Loja 21 - Flamengo',       'Zona Sul Loja 21 - Flamengo'),
    ('Zona Sul Loja 22 - S. Conrado',     'Zona Sul Loja 22 - S.Conrado'),
    ('Zona Sul Loja 23 - Barra',          'Zona Sul Loja 23 - Barra'),
    ('Zona Sul Loja 26 - Copacabana',     'Zona Sul Loja 26 - Copacabana'),
    ('Zona Sul Loja 27 - Ipanema',        'Zona Sul Loja 27 - Ipanema'),
    ('Zona Sul Loja 28 - Urca',           'Zona Sul Loja 28 - Urca'),
    ('Zona Sul Loja 29 - Flamengo',       'Zona Sul Loja 29 - Flamengo'),
    ('Zona Sul Loja 31 - Jd. Botânico',   'Zona Sul Loja 31 - Jd.Botânico'),
    ('Zona Sul Loja 32 - Laranjeiras',    'Zona Sul Loja 32 - Laranjeiras'),
    ('Zona Sul Loja 34 - Barra',          'Zona Sul Loja 34 - Barra'),
    ('Zona Sul Loja 35 - Barra',          'Zona Sul Loja 35 - Barra'),
    ('Zona Sul Loja 38 - Copacabana',     'Zona Sul Loja 38 - Copacabana'),
    ('Zona Sul Loja 40 - Ipanema',        'Zona Sul Loja 40 - Ipanema'),
    ('Zona Sul Loja 42 - Botafogo',       'Zona Sul Loja 42 - Botafogo'),
    ('Zona Sul Loja 44 - Barra',          'Zona Sul Loja 44 - Barra'),
    ('Zona Sul Loja 45 - Flamengo',       'Zona Sul Loja 45 - Flamengo'),
    ('Zona Sul Loja 46 - Botafogo',       'Zona Sul Loja 46 - Botafogo'),
    ('Zona Sul Loja 47',                  'Zona Sul Loja 47 - Catete'),
    ('Zona Sul Loja 48 - Recreio',        'Zona Sul Loja 48 - Recreio'),
    ('Zona Sul Loja 1129 - Olaria',       'Zona Sul Loja 1129 - Olaria'),
]
# Filiais sistema sem par no manual
sis_only_keys = [
    'Zona Sul Loja 06 - Gávea',
    'Zona Sul Loja 25 - Jd. Botânico',
    'Zona Sul Loja 30 - Laranjeiras',
    'Zona Sul Loja 33 - Humaií',
    'Zona Sul Loja 36 - Botafogo',
    'Zona Sul Loja 43 - Barra (Península)',
    'Zona Sul - Entrega Extra',
]
man_only_keys = [
    ('Zona Sul Loja 06 - Gávea',          'Zona Sul Loja 06 - Gávea'),
    ('Zona Sul Loja 25 - Jd.Botânico',    'Zona Sul Loja 25 - Jd.Botânico'),
    ('Zona Sul Loja 30 - Laranjeiras',          'Zona Sul Loja 30 - Laranjeiras'),
    ('Zona Sul Loja 33 - Humaié',          'Zona Sul Loja 33 - Humaié'),
    ('Zona Sul Loja 36 - Botafogo',             'Zona Sul Loja 36 - Botafogo'),
    ('Zona Sul Loja 43 - Barra (Península)','Zona Sul Loja 43 - Barra (Península)'),
]

divs = []

def add_div(filial, tipo, campo, val_sis, val_man, obs, row_man=None):
    label = filial
    if row_man: label += f' [man_R{row_man}]'
    divs.append({'filial': label, 'tipo': tipo, 'campo': campo,
                 'sis': val_sis, 'man': val_man, 'obs': obs})

def compare(sk, mk):
    # resolve chaves reais
    sd = None
    for k in sis:
        if k == sk: sd = sis[k]; break
    if sd is None:
        # partial
        for k in sis:
            if sk.lower() in k.lower() or k.lower() in sk.lower():
                sd = sis[k]; break
    if sd is None:
        return

    entries = []
    for k in man:
        if k == mk: entries = man[k]; break
    if not entries:
        for k in man:
            if mk.lower() in k.lower() or k.lower() in mk.lower():
                entries = man[k]; break

    if not entries:
        add_div(sk, 'FILIAL_AUSENTE_MANUAL', 'Filial', '-', '-',
                'Presente no sistema, ausente no manual')
        return

    for e in entries:
        row_tag = e['row'] if len(entries) > 1 else None

        # STATUS
        if sd['status'] == 'SEM RASTREADOR' and e['status'] not in ('SEM RASTREADOR', 'SEM DADOS'):
            add_div(sk, 'STATUS', 'Status', sd['status'], e['status'],
                    'Sistema sem rastreador, manual tem dados', row_tag)
        elif sd['status'] == 'COM GPS' and e['status'] == 'SEM RASTREADOR':
            add_div(sk, 'STATUS', 'Status', sd['status'], e['status'],
                    'Sistema com GPS, manual sem rastreador', row_tag)
        elif sd['status'] == 'SEM DADOS' and e['status'] == 'COM DADOS':
            add_div(sk, 'STATUS', 'Status', sd['status'], e['status'],
                    'Sistema sem dados GPS, manual tem dados', row_tag)

        # MOTORISTA - comparacao flexível
        mot_s = sd['mot'].upper().strip()
        mot_m = e['mot'].upper().strip()
        # pega primeiros tokens para comparar
        tokens_s = set(mot_s.replace('/', ' ').split())
        tokens_m = set(mot_m.replace('/', ' ').split())
        # Se nenhum token em comum, é divergencia
        common = tokens_s & tokens_m
        # Remove tokens muito curtos
        common = {t for t in common if len(t) > 2}
        if not common and mot_s and mot_m and mot_s != mot_m:
            add_div(sk, 'MOTORISTA', 'Motorista', sd['mot'], e['mot'],
                    'Motorista diferente', row_tag)

        # PLACA
        p_s = sd['placa'].replace('-','').upper()
        p_m = e['placa'].replace('-','').upper()
        if p_m and p_s and p_s != p_m:
            add_div(sk, 'PLACA', 'Placa', sd['placa'], e['placa'],
                    'Placa diferente', row_tag)

        # SAIDA CD
        scd_s = to_min(sd['saida_cd'])
        scd_m = to_min(e['saida_cd'])
        if scd_s is not None and scd_m is not None:
            diff = abs(scd_s - scd_m)
            if diff > 5:
                add_div(sk, 'SAIDA_CD', 'Saida CD',
                        sd['saida_cd'], e['saida_cd'],
                        f'Diferenca {diff} min', row_tag)

        # CHD
        chd_s = to_min(sd['chd'])
        chd_m = to_min(e['chd'])
        if chd_s is not None and chd_m is not None:
            diff = abs(chd_s - chd_m)
            if diff > 5:
                add_div(sk, 'CHD', 'CHD Loja',
                        sd['chd'], e['chd'],
                        f'Diferenca {diff} min', row_tag)

        # TEMPO
        t_s = to_min(sd['tempo'])
        t_m = to_min(e['tempo'])
        if t_s is not None and t_m is not None:
            diff = abs(t_s - t_m)
            if diff > 5:
                add_div(sk, 'TEMPO', 'Tempo em Loja',
                        sd['tempo'], e['tempo'],
                        f'Diferenca {diff} min', row_tag)

for sk, mk in pairs:
    compare(sk, mk)

# ================== RELATORIO ==================
print()
print('=' * 90)
print('RELATORIO DE DIVERGENCIAS - KPI ZONA SUL - 20/05/2026')
print('=' * 90)
print()
print('FONTES:')
print('  [SIS] KPI-ZONA_SUL-2026-05-20 (1).xlsx (aba 20.05)')
print('  [MAN] KPI ZONA SUL.xlsx (aba 20)')
print('  [ESC] ESCALA ZONA SUL - MAIO (7).xlsx')
print()

# 1. STATUS
print('=== 1. DIVERGENCIAS DE STATUS (SEM RASTREADOR vs COM GPS) ===')
subset = [d for d in divs if d['tipo'] == 'STATUS']
if subset:
    for d in subset:
        print(f'  Filial:  {d["filial"]}')
        print(f'  Sistema: {d["sis"]}')
        print(f'  Manual:  {d["man"]}')
        print(f'  Obs:     {d["obs"]}')
        print()
else:
    print('  Nenhuma divergência de status.\n')

# 2. MOTORISTA
print('=== 2. DIVERGENCIAS DE MOTORISTA ===')
subset = [d for d in divs if d['tipo'] == 'MOTORISTA']
if subset:
    for d in subset:
        print(f'  Filial:  {d["filial"]}')
        print(f'  Sistema: {d["sis"]}')
        print(f'  Manual:  {d["man"]}')
        print(f'  Obs:     {d["obs"]}')
        print()
else:
    print('  Nenhuma divergência de motorista.\n')

# 3. PLACA
print('=== 3. DIVERGENCIAS DE PLACA ===')
subset = [d for d in divs if d['tipo'] == 'PLACA']
if subset:
    for d in subset:
        print(f'  Filial:  {d["filial"]}')
        print(f'  Sistema: {d["sis"]}')
        print(f'  Manual:  {d["man"]}')
        print(f'  Obs:     {d["obs"]}')
        print()
else:
    print('  Nenhuma divergência de placa.\n')

# 4. SAIDA CD
print('=== 4. DIVERGENCIAS DE SAIDA DO CD (> 5 min) ===')
subset = [d for d in divs if d['tipo'] == 'SAIDA_CD']
if subset:
    for d in subset:
        print(f'  Filial:  {d["filial"]}')
        print(f'  Sistema: {d["sis"]}')
        print(f'  Manual:  {d["man"]}')
        print(f'  Obs:     {d["obs"]}')
        print()
else:
    print('  Nenhuma divergência de saída do CD.\n')

# 5. CHD
print('=== 5. DIVERGENCIAS DE CHD (Chegada na Loja) (> 5 min) ===')
subset = [d for d in divs if d['tipo'] == 'CHD']
if subset:
    for d in subset:
        print(f'  Filial:  {d["filial"]}')
        print(f'  Sistema: {d["sis"]}')
        print(f'  Manual:  {d["man"]}')
        print(f'  Obs:     {d["obs"]}')
        print()
else:
    print('  Nenhuma divergência de CHD.\n')

# 6. TEMPO
print('=== 6. DIVERGENCIAS DE TEMPO EM LOJA (> 5 min) ===')
subset = [d for d in divs if d['tipo'] == 'TEMPO']
if subset:
    for d in subset:
        print(f'  Filial:  {d["filial"]}')
        print(f'  Sistema: {d["sis"]}')
        print(f'  Manual:  {d["man"]}')
        print(f'  Obs:     {d["obs"]}')
        print()
else:
    print('  Nenhuma divergência de tempo em loja.\n')

print()
print(f'TOTAL DE DIVERGENCIAS: {len(divs)}')
print()
# Sumarizar por tipo
tipos = ['STATUS','MOTORISTA','PLACA','SAIDA_CD','CHD','TEMPO']
for t in tipos:
    n = len([d for d in divs if d['tipo'] == t])
    if n: print(f'  {t}: {n}')

# =================== ANALISE DE PADROES ===================
print()
print('=' * 90)
print('ANALISE DE PADROES CRITICOS (conforme dias 18/19)')
print('=' * 90)

print()
print('--- P1: GPS de madrugada para loja que deveria ser atendida somente a tarde ---')
# Lojas tipicamente da tarde (a partir de 12h): 28, 29, 12, 09, 15, 38, 03
tarde = {
    'Zona Sul Loja 03 - Copacabana I': 12,
    'Zona Sul Loja 09 - Ipanema': 12,
    'Zona Sul Loja 12 - Leme': 12,
    'Zona Sul Loja 15 - Leblon': 12,
    'Zona Sul Loja 28 - Urca': 12,
    'Zona Sul Loja 29 - Flamengo': 12,
    'Zona Sul Loja 38 - Copacabana': 12,
    'Zona Sul Loja 21 - Flamengo': 12,
    'MEGA BOX 01 - Olaria': 12,
}
for sk, expected_hour in tarde.items():
    if sk in sis:
        sd = sis[sk]
        m = to_min(sd['saida_cd'])
        if m is not None and m < expected_hour * 60:
            print(f'  ALERTA: {sk}')
            print(f'    Saida CD sistema: {sd["saida_cd"]} (esperado apos {expected_hour}:00)')
            print(f'    Motorista: {sd["mot"]} | Placa: {sd["placa"]}')
            print()

print()
print('--- P2: Entrega dupla perdida (sistema registra 1 de 2 entregas na mesma filial) ---')
# Filiais com 2+ entradas no manual
for mk, entries in man.items():
    if len(entries) >= 2:
        # verificar se todas aparecem no sistema
        print(f'  MANUAL tem {len(entries)} entregas para: {mk}')
        for e in entries:
            print(f'    R{e["row"]}: mot={e["mot"]} | placa={e["placa"]} | saida_cd={e["saida_cd"]} | chd={e["chd"]} | tempo={e["tempo"]} | status={e["status"]}')
        # verificar o que o sistema tem
        for sk in sis:
            if mk.lower() in sk.lower() or sk.lower() in mk.lower():
                sd = sis[sk]
                print(f'    SISTEMA: mot={sd["mot"]} | placa={sd["placa"]} | saida_cd={sd["saida_cd"]} | chd={sd["chd"]} | tempo={sd["tempo"]}')
        print()

print()
print('--- P3: Tempo absurdo de permanencia (>2h no sistema, <1h no manual) ---')
for sk, mk in pairs:
    sd = None
    for k in sis:
        if k == sk: sd = sis[k]; break
    if not sd: continue
    entries = []
    for k in man:
        if k == mk: entries = man[k]; break
    if not entries: continue
    t_s = to_min(sd['tempo'])
    for e in entries:
        t_m = to_min(e['tempo'])
        if t_s and t_s > 120 and t_m and t_m < 60:
            print(f'  Filial: {sk}')
            print(f'    Sistema: {sd["tempo"]} ({t_s} min) | Manual: {e["tempo"]} ({t_m} min)')
            print(f'    Motorista sistema: {sd["mot"]} | Placa: {sd["placa"]}')
            print()

print()
print('--- P4: Veiculo/motorista errado atribuido a filial ---')
# Ja capturado nas divergencias de MOTORISTA e PLACA
subset = [d for d in divs if d['tipo'] in ('MOTORISTA', 'PLACA')]
if subset:
    for d in subset:
        print(f'  Filial: {d["filial"]}')
        print(f'    Campo: {d["campo"]}')
        print(f'    Sistema: {d["sis"]}')
        print(f'    Manual:  {d["man"]}')
else:
    print('  Nenhuma detectada.')
print()

print()
print('--- P5: CHD impossivel (1-2 minutos apos saida do CD) ---')
for sk in sis:
    sd = sis[sk]
    scd = to_min(sd['saida_cd'])
    chd = to_min(sd['chd'])
    if scd and chd:
        diff = chd - scd
        if 0 < diff <= 5:
            print(f'  Filial: {sk}')
            print(f'    Saida CD: {sd["saida_cd"]} | CHD: {sd["chd"]} | Diferenca: {diff} min')
            print(f'    Motorista: {sd["mot"]} | Placa: {sd["placa"]}')
            print()

print()
print('--- P6: Dados GPS duplicados para duas filiais (mesmo horario em filiais distintas) ---')
# Verificar pares de filiais com mesmos horarios
items = [(sk, sd) for sk, sd in sis.items() if sd['status'] == 'COM GPS']
for i in range(len(items)):
    for j in range(i+1, len(items)):
        ska, sda = items[i]
        skb, sdb = items[j]
        # Mesmo veiculo
        if sda['placa'] != sdb['placa'] or not sda['placa']: continue
        # Mesmos horarios
        if sda['saida_cd'] == sdb['saida_cd'] and sda['chd'] == sdb['chd'] and sda['saida_loja'] == sdb['saida_loja']:
            print(f'  GPS DUPLICADO - Placa {sda["placa"]}:')
            print(f'    Filial A: {ska} | saida_cd={sda["saida_cd"]} | chd={sda["chd"]} | saida_loja={sda["saida_loja"]}')
            print(f'    Filial B: {skb} | saida_cd={sdb["saida_cd"]} | chd={sdb["chd"]} | saida_loja={sdb["saida_loja"]}')
            print()

print()
print('--- OBSERVACOES ADICIONAIS ---')
print()
# Loja 21 - sistema registra motorista/placa da madrugada (SIDNEI/LQE-5401) mas manual tem WILLIAM/LTQ-0783 turno manha
sk = 'Zona Sul Loja 21 - Flamengo'
if sk in sis:
    sd = sis[sk]
    entries = man.get('Zona Sul Loja 21 - Flamengo', [])
    print(f'  {sk}:')
    print(f'    Sistema: mot={sd["mot"]} | placa={sd["placa"]} | saida_cd={sd["saida_cd"]} | chd={sd["chd"]} | tempo={sd["tempo"]}')
    for e in entries:
        print(f'    Manual R{e["row"]}: mot={e["mot"]} | placa={e["placa"]} | saida_cd={e["saida_cd"]} | chd={e["chd"]} | tempo={e["tempo"]}')
print()

# Loja 09 e 15 e 27 - sistema atribui RODRIGO/VLADIMIR KQY-9E24 com mesmo horario
for loja in ['Zona Sul Loja 09 - Ipanema', 'Zona Sul Loja 15 - Leblon', 'Zona Sul Loja 27 - Ipanema']:
    if loja in sis:
        sd = sis[loja]
        print(f'  {loja}: mot={sd["mot"]} | placa={sd["placa"]} | saida_cd={sd["saida_cd"]} | chd={sd["chd"]} | saida_loja={sd["saida_loja"]} | tempo={sd["tempo"]}')

print()
# Loja 33, 36, 45 - mesmo horario GPS
for loja in ['Zona Sul Loja 33 - Humaií', 'Zona Sul Loja 36 - Botafogo', 'Zona Sul Loja 45 - Flamengo']:
    for k in sis:
        if '33' in k and 'Humai' in k or '36' in k and 'Botafogo' in k or '45' in k and 'Flamengo' in k:
            sd = sis[k]
            if loja[:15] in k:
                print(f'  {k}: mot={sd["mot"]} | placa={sd["placa"]} | saida_cd={sd["saida_cd"]} | chd={sd["chd"]} | saida_loja={sd["saida_loja"]} | tempo={sd["tempo"]}')
